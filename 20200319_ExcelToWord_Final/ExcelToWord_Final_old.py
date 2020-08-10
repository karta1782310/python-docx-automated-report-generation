#!/usr/bin/env python
# coding: utf-8

# In[1]:


import os
import copy
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm

from math import ceil
from openpyxl import load_workbook
from MyPythonDocx import *


# In[2]:


def prevent_document_break(document):
    tags = document.element.xpath('//w:tr/w:tc/w:tbl/w:tr/w:tc/w:p')
    rows = len(tags)
    for row in range(0, rows):
        tag = tags[row]  # Specify which <w:r> tag you want
        child = OxmlElement('w:cantSplit')  # Create arbitrary tag
        tag.append(child)  # Append in the new tag


# In[3]:


def plt_bar(title, names, values, picture):
    plt.style.use('seaborn-whitegrid')

    plt.figure(figsize=(9, 4))
    plt.title(title, size=22)
    plt.ylim(0, (max(values) * 6/5 + 20) // 10 * 10)

    plt.rcParams['font.sans-serif'] = ['DFKai-SB']  # 替換sans-serif字型
    plt.rcParams['axes.unicode_minus'] = False  # 解決座標軸負數的負號顯示問題
    plt.rcParams.update({'font.size': 16})
    plt.xticks(size=22)

    bar = plt.bar(names, values, color=[
                  'cornflowerblue', 'indianred', 'yellowgreen'])
    for rect in bar:
        h = rect.get_height()
        plt.text(rect.get_x()+rect.get_width()/2.0, h, "%d\n" %
                 h, ha='center', va='bottom')

    plt.savefig(picture, dpi=300)
    # plt.show()
    
def sort_ip(ips):
    return sorted([i.strip() for i in ips], key=lambda x: int(''.join(
        (lambda a: lambda v: a(a, v))(lambda s, x: x if len(x) == 3 else s(s, '0'+x))(i) for i in x.split('.'))))


# In[4]:


def doc_risk_bar(doc, data, ips=[]):

    names = ['高風險', '中風險', '低風險']
    values = [0, 0, 0]
    level = ""
    tmp_ips = copy.deepcopy(ips)

    # calculate needed data
    for row in data:
        if tmp_ips:
            for ip in tmp_ips:
                if row[0] == ip:
                    level = row[7]
        else:
            level = row[7]

        if level == '高':
            values[0] += 1
        elif level == "中":
            values[1] += 1
        elif level == "低":
            values[2] += 1

        level = ""

    # find the mark
    for table in doc.tables:
        if "riskBar" in table.cell(0, 0).paragraphs[0].text:
            table.allow_autofit = True

            # add pic in table
            plt_bar('風險', names, values, "risk_bar.png")

            main_cell = table.cell(0, 0)
            main_cell.paragraphs[0].text = ""

            parag = main_cell.paragraphs[0]
            parag.add_run().add_picture("risk_bar.png", width=Cm(15))
            parag.paragraph_format.alignment = WD_ALIGN_PARAGRAPH.CENTER

    os.remove("risk_bar.png")

    # 替換標定資料
    for parag in doc.paragraphs:
        row = parag.text
        if "weaknessCount" in row or "highRiskWeaknessCount" in row or "middleRiskWeaknessCount" in row or "lowRiskWeaknessCount" in row:
            row = row.replace("weaknessCount", str(
                values[0]+values[1]+values[2]), 1)
            row = row.replace("highRiskWeaknessCount", str(values[0]), 1)
            row = row.replace("middleRiskWeaknessCount", str(values[1]), 1)
            row = row.replace("lowRiskWeaknessCount", str(values[2]), 1)
            parag.text = row


# In[5]:


# In[6]:

def doc_ip_list(doc, data, ips=[]):

    tmp_ips = copy.deepcopy(ips)

    # calculate needed data
    if not tmp_ips:
        for row in data:
            if row[0] not in tmp_ips and row[0]:
                tmp_ips.append(row[0])

        del(tmp_ips[0])  # delete "IP初始"
        tmp_ips = sort_ip(tmp_ips)

    # find the mark
    for table in doc.tables:
        if "ipList" in table.cell(0, 0).paragraphs[0].text:

            # add list table
            main_cell = table.cell(0, 0)
            Parag.delete(main_cell.paragraphs[0])

            cols, ip_cnt = 0, len(tmp_ips)
            lines = 20 if len(tmp_ips) < 4*20 else ceil(len(tmp_ips)/4)

            cols = ceil(ip_cnt / lines)
            rows = ceil(ip_cnt / cols)

            table = main_cell.add_table(1+rows, cols)
            table.style = 'Table Web 1'
            table.allow_autofit = True
            table.alignment = WD_TABLE_ALIGNMENT.CENTER

            cell = table.cell(0, 0).merge(table.cell(0, cols-1))
            cell.text = "IP 清單"
            Table.set_cell_color(cell, GRAY)

            for i, ip in enumerate(tmp_ips):
                idx = i // rows
                table.cell(1 + (i % rows), idx).text = ip

            Table.set_content_font(table, size=Pt(
                12), align=1 if ip_cnt < 2*20 else 0)
            Parag.set_font(cell.paragraphs, bold=True, align=1)

            # title = doc.tables[1].cell(0, 0).paragraphs[0]
            # title.text = "表3-1: 主機弱點掃描網段清單"
            # Parag.set_font(title)


# In[6]:



def doc_high_risk_ip_list(doc, data, ips=[]):

    tmp_ips = copy.deepcopy(ips)
    empty = True

    # calculate needed data
    if not tmp_ips:
        for row in data:
            ip, risk = row[0], row[7]
            if ip not in tmp_ips and risk == "高":
                tmp_ips.append(ip)

        if tmp_ips:
            empty = False

            # sort ip
            tmp_ips = sort_ip(tmp_ips)
    else:
        for row in data:
            ip, risk = row[0], row[7]
            if ip == tmp_ips[0] and risk == "高":
                empty = False

    # find the mark
    for table in doc.tables:
        if "highRiskIpList" in table.cell(0, 0).paragraphs[0].text:
            if not empty:
                # add list table
                main_cell = table.cell(0, 0)
                Parag.delete(main_cell.paragraphs[0])

                table = main_cell.add_table(2, 4)
                table.style = 'Table Web 1'
                cell = table.cell(0, 0).merge(table.cell(0, 3))
                cell.text = "高風險主機 IP 清單"
                Table.set_cell_color(cell, GRAY)

                for i in range(4):
                    Parag.delete(table.cell(1, i).paragraphs[0])

                table.cell(1, 3).add_paragraph(
                    "根據國際標準CVSS 第三版風險評比所述，左列IP屬於高風險設備，基本上容易被入侵。")

                for i, ip in enumerate(sorted(tmp_ips)):
                    idx = i // ceil(len(tmp_ips)/3)
                    table.cell(1, idx).add_paragraph(ip)

                Table.set_content_font(table, align=0)
                Parag.set_font(cell.paragraphs, bold=True)

            else:
                table.cell(0, 0).paragraphs[0].text = "無"

                
    # 替換標定資料
    for parag in doc.paragraphs:
        if "highRiskIpCount" in parag.text:
            parag.text = parag.text.replace(
                "highRiskIpCount", str(len(tmp_ips)), 1)


# In[8]:


def doc_cve_describe(doc, data, ips=[], risk='高'):

    dic, describe, describe2 = {}, [], []
    level, idx = "", 0
    tmp_ips = copy.deepcopy(ips)

    cve_cnt = []
    
    # calculate needed data
    for row in data:
        if tmp_ips:
            for ip in tmp_ips:
                if row[0] == ip:
                    level = row[7]
        else:
            level = row[7]

        if level == risk:

            # 如果弱點沒出現過，則記在字典
            if row[8] not in dic:
                dic.setdefault(row[8], idx)
                describe.append([idx+1, "立即的", row[7]])
                describe2.append([row[8]+"\n", row[9], row[10]+"\n", [row[0]]])
                cve_cnt.append([1, row[8]])
                idx += 1

            # 如果弱點出現過，且依據的 IP 沒存在 describe2 中，則新增進去
            elif row[0] not in describe2[dic[row[8]]][3]:
                describe2[dic[row[8]]][3].append(row[0])
                
                # 計算出現過幾次
                cve_cnt[dic[row[8]]][0] = cve_cnt[dic[row[8]]][0] + 1
                

        level = ""  # reset

    #  High Risk Top 10
    cve_cnt.sort(reverse=True)
        
    # sort ip
    for c in range(idx):
        describe2[c][3] = sort_ip(describe2[c][3])

    # find the mark
    for table in doc.tables:
        if "cveDescribe" in table.cell(0, 0).paragraphs[0].text:
            if idx:
                # add describe table
                main_cell = table.cell(0, 0)
                Parag.delete(main_cell.paragraphs[0])

                for cnt in range(idx):

                    des_table = main_cell.add_table(10, 3)
                    des_table.style = "Table Web 1"
                    des_table.allow_autofit = True
                    des_table.alignment = WD_TABLE_ALIGNMENT.CENTER
                    Table.col_widths(des_table, 5.2, 5.2, 5.2)

                    for i in range(7):
                        des_table.cell(i+2, 0).merge(des_table.cell(i+2, 2))

                    tag = ['序號', "急迫性", "風險性"]
                    for i, title in enumerate(tag):
                        cell = des_table.cell(0, i)
                        cell.text = title
                        Table.set_cell_color(cell, GRAY)
                        Parag.set_font(cell.paragraphs, bold=True)

                        cell2 = des_table.cell(1, i)
                        cell2.text = str(describe[cnt][i])
                        cell2.vertical_alignment = 1
                        Parag.set_font(cell2.paragraphs)

                    tag2 = ["弱點名稱", "弱點說明", "改善建議", "受影響 IP 位置"]
                    for i, title in enumerate(tag2):
                        cell = des_table.cell(2*(i+1), 0)
                        cell.text = title
                        Table.set_cell_color(cell, GRAY)
                        Parag.set_font(cell.paragraphs, bold=True)

                        if i < 3:
                            cell2 = des_table.cell(2*(i+1)+1, 0)
                            cell2.text = str(describe2[cnt][i])
                            Parag.set_font(cell2.paragraphs[0], align=0)
                        else:
                            row_cnt = ceil(len(describe2[cnt][3]) / 3)
                            for j, ip in enumerate(describe2[cnt][3]):
                                idx = j // row_cnt
                                cell2 = des_table.cell(2*(i+1)+1, idx)
                                cell2.paragraphs[0].add_run(
                                    describe2[cnt][3][j])

                                # 若不是每一欄最後一個ip，則加上換行
                                if j % row_cnt + 1 < row_cnt:
                                    cell2.paragraphs[0].add_run("\n")

                            for j in range(0, 3):
                                Parag.set_font(des_table.cell(
                                    9, j).paragraphs, align=WD_ALIGN_PARAGRAPH.LEFT)
            else:
                table.cell(0, 0).paragraphs[0].text = "無"


    for parag in doc.paragraphs:
        if 'highRiskCnt' in parag.text:
            parag.text = parag.text.replace('highRiskCnt', str(len(cve_cnt)), 1)
    
        if 'highRiskTop10' in parag.text:
            title = ['Top 10', '弱點名稱', '弱點數']
            
            table = doc.add_table(11, 3, style='Table Web 1')            
            for i in range(11):
                for j in range(3):
                    cell = table.cell(i, j)
                    cell.vertical_alignment = 1
                    if not i:
                        cell.text = title[j]
                        Parag.set_font(cell.paragraphs, bold=True)
                        Table.set_cell_color(cell, GRAY)
                    else:
                        if j == 0:
                            cell.text = str(i)
                            Parag.set_font(cell.paragraphs)
                        elif j == 1:
                            cell.text = cve_cnt[i-1][1]
                            Parag.set_font(cell.paragraphs, align=0)
                        elif j == 2:
                            cell.text = str(cve_cnt[i-1][0])
                            Parag.set_font(cell.paragraphs)
                                        
                    
            Table.col_widths(table, 1, 4.5, 1)
            
            Parag.insert_paragraph_after(parag).add_run().add_break(WD_BREAK.PAGE)
            Table.move_table_after(table, parag)
            Parag.delete(parag)


# In[9]:


def ExcelToWord(excel, sheet, word, company, date, ips, save):

    wb = load_workbook(excel)
    data = list(wb[sheet].values)
    
    doc = Document(word)
    
    doc_ip_list(doc, data, ips)
    doc_risk_bar(doc, data, ips)
    doc_high_risk_ip_list(doc, data, ips)
    doc_cve_describe(doc, data, ips)

    for parag in doc.paragraphs:
        row = parag.text
        if "YYYY" in row or "MMMM" in row or "DDDD" in row:
            row = row.replace("YYYY", date[0], 1)
            row = row.replace("MMMM", date[1], 1)
            row = row.replace("DDDD", date[2], 1)
            parag.text = row
            Parag.set_font(parag, align=0)

        if "OOOOO" in row:
            tmp_str = row.split("OOOOO")
            parag.clear()
            parag.add_run(tmp_str[0])
            for i in range(1, len(tmp_str)):
                parag.add_run(company).underline = True
                parag.add_run(tmp_str[i])
                
        if "referExcel" in row:
            tmp_str = excel.split('/')[-1]
            parag.text = row.replace('referExcel', tmp_str, 1)

    doc.save(save)


# In[10]:


# if __name__ == '__main__':
#     excel = '/Users/SystexB/Documents/公司資料/初掃 xlsx/嘉碩初掃報告-0316.xlsx'
#     sheet = 'All'
#     word = '/Users/SystexB/Documents/公司資料/proto2.docx'
#     company = '嘉頓'
#     date = ['2020', '03', '19']
#     save = '/Users/SystexB/Desktop/demo2_嘉碩初掃報告-0316.docx'
    
#     main(excel, sheet, word, company, date, [], save)


# In[ ]:




