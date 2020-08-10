#!/usr/bin/env python
# coding: utf-8
# version:03/25

import os
import copy
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm

from math import ceil
from openpyxl import load_workbook
from pandas.core.frame import DataFrame
from MyPythonDocx import *

# 產生柱狀圖並儲存在當前目錄
def plt_bar(title, names, values, picture):
    plt.style.use('seaborn-whitegrid')

    plt.figure(figsize=(9, 4))
    plt.title(title, size=22)
    plt.ylim(0, (max(values) * 6/5 + 20) // 10 * 10)

    plt.rcParams['font.sans-serif'] = ['DFKai-SB']  # 替換sans-serif字型
    plt.rcParams['axes.unicode_minus'] = False  # 解決座標軸負數的負號顯示問題
    plt.rcParams.update({'font.size': 16})
    plt.xticks(size=22)

    bar = plt.bar(names, values, color=[
                  'cornflowerblue', 'indianred', 'yellowgreen'])
    for rect in bar:
        h = rect.get_height()
        plt.text(rect.get_x()+rect.get_width()/2.0, h, "%d\n" %
                 h, ha='center', va='bottom')

    plt.savefig(picture, dpi=300)
    # plt.show()

# 排序 ip list
def sort_ip(ips):
    return sorted([i.strip() for i in ips], key=lambda x: int(''.join(
        (lambda a: lambda v: a(a, v))(lambda s, x: x if len(x) == 3 else s(s, '0'+x))(i) for i in x.split('.'))))

# 計算高中低風險個數
def cal_risk_cnt(page, ips=[]):

    values = []
    
    df = DataFrame(page[1:], columns=page[0])
    
    opt = (df['IP位址'] == ips[0]) if ips else True 
    mask = (df['嚴重程度'] == '高')
    mask1 = (df['嚴重程度'] == '中')
    mask2 = (df['嚴重程度'] == '低')
    
    return [df[mask & opt].shape[0], df[mask1 & opt].shape[0], df[mask2 & opt].shape[0]]

# 提取全部的 ip 清單
def cal_ips(wb, sheet, ips=[]):
    
    if ips:
        return ips[0]
    
    data = None
    try:
        data = list(wb['主機'].values)
    except:
        data = list(wb[sheet].values)

    df = DataFrame(data[1:], columns=data[0])
    df_ips = list(set(df['IP位址']))
    df_ips = sort_ip(df_ips)
    
    return df_ips

# 提取指定風險的 ip 清單
def cal_high_risk_ips(data, risk='高', ips=[]):
    
    if ips:
        return ips[0]
    
    df = DataFrame(data[1:], columns=data[0])    
    mask = df['嚴重程度'] == risk
    df_ips = list(set(df[mask]['IP位址']))
    df_ips = sort_ip(df_ips)
    
    return df_ips

# 產出指定 ip 的風險描述，包括序號、急迫性、風險性 (describe)，名稱、說明、建議、影響IP (describe2)
def cve_detail(data, ips, risk):

    dic, describe, describe2, cve_cnt = {}, [], [], []
    
    level, idx = "", 0
    tmp_ips = copy.deepcopy(ips)
    
    # calculate needed data
    for row in data:
        if tmp_ips:
            for ip in tmp_ips:
                if row[0] == ip:
                    level = row[7]
        else:
            level = row[7]

        if level == risk:
            risk_name = row[8]
            risk_ip = row[0]
            # 如果弱點沒出現過，則記在字典
            if risk_name not in dic:
                dic.setdefault(row[8], idx)
                # 序號, 急迫性, 風險性
                describe.append([idx+1, "立即的", row[7]])
                # 弱點名稱, 弱點說明, 改善建議, 受影響 IP 位置
                describe2.append([risk_name+"\n", row[9], row[10]+"\n", [risk_ip]]) # 
                cve_cnt.append([1, risk_name])
                idx += 1

            # 如果弱點出現過，且依據的 IP 沒存在 describe2 中，則新增進去
            elif row[0] not in describe2[dic[risk_name]][3]:
                describe2[dic[risk_name]][3].append(risk_ip)
                
                # 計算出現過幾次
                cve_cnt[dic[risk_name]][0] = cve_cnt[dic[risk_name]][0] + 1                

        level = ""  # reset

    return dic, describe, describe2, cve_cnt

# 有 doc 前綴的 funcation，都是將資料寫入 docx 中

def doc_risk_bar(doc, data, ips=[]):

    names = ['高風險', '中風險', '低風險']

    # calculate needed data
    values = cal_risk_cnt(data, ips)

    # find the mark
    for table in doc.tables:
        if "riskBar" in table.cell(0, 0).paragraphs[0].text:
            # table.allow_autofit = True

            # add pic in table
            plt_bar('風險', names, values, "risk_bar.png")

            main_cell = table.cell(0, 0)
            main_cell.paragraphs[0].text = ""

            parag = main_cell.paragraphs[0]
            parag.add_run().add_picture("risk_bar.png", width=Cm(15))
            parag.paragraph_format.alignment = WD_ALIGN_PARAGRAPH.CENTER
            os.remove("risk_bar.png")
    
    # 替換標定資料
    for parag in doc.paragraphs:
        row = parag.text
        if "weaknessCount" in row or "highRiskWeaknessCount" in row or "middleRiskWeaknessCount" in row or "lowRiskWeaknessCount" in row:
            row = row.replace("weaknessCount", str(
                values[0]+values[1]+values[2]), 1)
            row = row.replace("highRiskWeaknessCount", str(values[0]), 1)
            row = row.replace("middleRiskWeaknessCount", str(values[1]), 1)
            row = row.replace("lowRiskWeaknessCount", str(values[2]), 1)
            parag.text = row
            


def doc_ip_list(doc, wb, sheet, ips=[]):

    # calculate needed data
    tmp_ips = cal_ips(wb, sheet, ips)

    # find the mark
    for table in doc.tables:
        if "ipList" in table.cell(0, 0).paragraphs[0].text:

            # add list table
            main_cell = table.cell(0, 0)
            Parag.delete(main_cell.paragraphs[0])

            cols, ip_cnt = 0, len(tmp_ips)
            lines = 20 if len(tmp_ips) < 4*20 else ceil(len(tmp_ips)/4)

            cols = ceil(ip_cnt / lines)
            rows = ceil(ip_cnt / cols)

            table = main_cell.add_table(1+rows, cols)
            table.style = 'Table Grid'
            table.allow_autofit = True
            table.alignment = WD_TABLE_ALIGNMENT.CENTER

            cell = table.cell(0, 0).merge(table.cell(0, cols-1))
            cell.text = "IP 清單"
            Table.set_cell_color(cell, GRAY)

            for i, ip in enumerate(tmp_ips):
                idx = i // rows
                table.cell(1 + (i % rows), idx).text = ip

            Table.set_content_font(table, size=Pt(12), align=1 if ip_cnt < 2*20 else 0)
            Parag.set_font(cell.paragraphs, bold=True, align=1)

    for parag in doc.paragraphs:
        if 'totalIpCnt' in parag.text:
            parag.text = parag.text.replace('totalIpCnt', str(len(tmp_ips)))

def doc_high_risk_ip_list(doc, data, ips=[]):

    # calculate needed data
    tmp_ips = cal_high_risk_ips(data, ips=ips)

    # find the mark
    for table in doc.tables:
        if "highRiskIpList" in table.cell(0, 0).paragraphs[0].text:
            if tmp_ips:
                # add list table
                main_cell = table.cell(0, 0)
                Parag.delete(main_cell.paragraphs[0])

                columns = len(tmp_ips)+1 if len(tmp_ips)<3 else 4
                table = main_cell.add_table(2, columns)
                table.style = 'Table Grid'
                cell = table.cell(0, 0).merge(table.cell(0, columns-1))
                cell.text = "高風險主機 IP 清單"
                Table.set_cell_color(cell, GRAY)

                for i in range(0, columns):
                    Parag.delete(table.cell(1, i).paragraphs[0])

                table.cell(1, columns-1).add_paragraph(
                    "根據國際標準CVSS 第三版風險評比所述，左列IP屬於高風險設備，基本上容易被入侵。")

                for i, ip in enumerate(sorted(tmp_ips)):
                    idx = i // ceil(len(tmp_ips)/3)
                    table.cell(1, idx).add_paragraph(ip)

                Table.set_content_font(table, align=0)
                Parag.set_font(cell.paragraphs, bold=True)

            else:
                table.cell(0, 0).paragraphs[0].text = "無高風險主機，請刪除本表格。"

                
    # 替換標定資料
    for parag in doc.paragraphs:
        if "highRiskIpCnt" in parag.text:
            parag.text = parag.text.replace(
                "highRiskIpCnt", str(len(tmp_ips)), 1)


def doc_cve_describe(doc, data, risk, ips=[]):

    dic, describe, describe2, cve_cnt = cve_detail(data, ips, risk)

    #  High Risk Top 10
    cve_cnt.sort(reverse=True)
            
    idx = len(cve_cnt)

    # sort ip
    for c in range(idx):
        describe2[c][3] = sort_ip(describe2[c][3])

    # find the mark
    for table in doc.tables:
        if "cveDescribe"+risk in table.cell(0, 0).paragraphs[0].text:
            if idx:
                # add describe table
                main_cell = table.cell(0, 0)
                Parag.delete(main_cell.paragraphs[0])

                for cnt in range(idx):

                    des_table = main_cell.add_table(10, 3)
                    des_table.style = 'Table Grid'
                    des_table.allow_autofit = True
                    des_table.alignment = WD_TABLE_ALIGNMENT.CENTER
#                     Table.col_widths(des_table, 5.2, 5.2, 5.2)

                    for i in range(7):
                        des_table.cell(i+2, 0).merge(des_table.cell(i+2, 2))

                    tag = ['序號', "急迫性", "風險性"]
                    for i, title in enumerate(tag):
                        cell = des_table.cell(0, i)
                        cell.text = title
                        Table.set_cell_color(cell, GRAY)
                        Parag.set_font(cell.paragraphs, bold=True)

                        cell2 = des_table.cell(1, i)
                        cell2.text = str(describe[cnt][i])
                        cell2.vertical_alignment = 1
                        Parag.set_font(cell2.paragraphs)

                    tag2 = ["弱點名稱", "弱點說明", "改善建議", "受影響 IP 位置"]
                    for i, title in enumerate(tag2):
                        cell = des_table.cell(2*(i+1), 0)
                        cell.text = title
                        Table.set_cell_color(cell, GRAY)
                        Parag.set_font(cell.paragraphs, bold=True)

                        if i < 3:
                            cell2 = des_table.cell(2*(i+1)+1, 0)
                            cell2.text = str(describe2[cnt][i])
                            Parag.set_font(cell2.paragraphs, align=0)
                        else:
                            row_cnt = ceil(len(describe2[cnt][3]) / 3)
                            for j, ip in enumerate(describe2[cnt][3]):
                                idx = j // row_cnt
                                cell2 = des_table.cell(2*(i+1)+1, idx)
                                cell2.paragraphs[0].add_run(describe2[cnt][3][j])

                                # 若不是每一欄最後一個ip，則加上換行
                                if j % row_cnt + 1 < row_cnt:
                                    cell2.paragraphs[0].add_run("\n")

                            for j in range(0, 3):
                                Parag.set_font(des_table.cell(9, j).paragraphs, align=0)
            else:
                table.cell(0, 0).paragraphs[0].text = "無"+risk+"風險, 請刪除本表格。"

        if 'highRiskTop10' in table.cell(0, 0).paragraphs[0].text and risk == '高':
            if len(cve_cnt):
                Parag.delete(table.cell(0, 0).paragraphs[0])
                title = ['Top 10', '弱點名稱', '弱點IP個數']

                max_row = 1+len(cve_cnt)  if len(cve_cnt) < 10 else 11
                table = table.cell(0, 0).add_table(max_row, 3)          
                table.style='Table Grid'
                for i in range(max_row):
                    for j in range(3):
                        cell = table.cell(i, j)
                        cell.vertical_alignment = 1
                        if not i:
                            cell.text = title[j]
                            Parag.set_font(cell.paragraphs, bold=True)
                            Table.set_cell_color(cell, GRAY)
                        else:
                            if j == 0:
                                cell.text = str(i)
                                Parag.set_font(cell.paragraphs)
                            elif j == 1:
                                cell.text = cve_cnt[i-1][1]
                                Parag.set_font(cell.paragraphs, align=0)
                            elif j == 2:
                                cell.text = str(cve_cnt[i-1][0])
                                Parag.set_font(cell.paragraphs)
                Table.col_widths(table, 1, 4.5, 1.5)

            else:
                table.cell(0, 0).paragraphs[0].text = '無高風險, 請刪除本表格。'

    for parag in doc.paragraphs:
        if 'cveCnt'+risk in parag.text:
            parag.text = parag.text.replace('cveCnt'+risk, str(len(cve_cnt)), 1)

def ExcelToWord(excel, sheet, word, company, date, ips, save):

    wb = load_workbook(excel)
    # wb = pd.read_excel(excel, sheet_name=None)
    data = list(wb[sheet].values)
    
    doc = Document(word)
    
    doc_ip_list(doc, wb, sheet, ips)
    doc_risk_bar(doc, data, ips)
    doc_high_risk_ip_list(doc, data, ips)
    doc_cve_describe(doc, data, '高', ips)
    doc_cve_describe(doc, data, '中', ips)
    doc_cve_describe(doc, data, '低', ips)

    for parag in doc.paragraphs:
        row = parag.text
        if "YYYY" in row or "MMMM" in row or "DDDD" in row:
            row = row.replace("YYYY", date[0], 1)
            row = row.replace("MMMM", date[1], 1)
            row = row.replace("DDDD", date[2], 1)
            parag.text = row
            Parag.set_font(parag, align=0)

        if "OOOOO" in row:
            tmp_str = row.split("OOOOO")
            parag.clear()
            parag.add_run(tmp_str[0])
            for i in range(1, len(tmp_str)):
                parag.add_run(company).underline = True
                parag.add_run(tmp_str[i])
                
        if "referExcel" in row:
            tmp_str = excel.split('/')[-1]
            parag.text = row.replace('referExcel', tmp_str, 1)

    doc.save(save)

if __name__ == '__main__':
    excel = '/Users/yucc/Desktop/report (1).xlsx'
    sheet = 'All'
    word = '/Volumes/DATA/GoodleDrive/Internship/公司資料/模板/proto_excel.docx'
    company = '翡翠'
    date = ['2020', '5', '4']
    save = '/Users/yucc/Desktop/test.docx'
    
    ExcelToWord(excel, sheet, word, company, date, [], save)
