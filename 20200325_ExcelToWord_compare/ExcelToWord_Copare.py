#!/usr/bin/env python
# coding: utf-8


import copy

from openpyxl import load_workbook
from openpyxl import Workbook

from pandas.core.frame import DataFrame
from MyPythonDocx import *


def cal_risk_cnt(page, ips=[]):

    values = []
    
    df = DataFrame(page[1:], columns=page[0])
    
    opt = (df['IP位址'] == ips[0]) if ips else True 
    mask = (df['嚴重程度'] == '高')
    mask1 = (df['嚴重程度'] == '中')
    mask2 = (df['嚴重程度'] == '低')
    
    return [df[mask & opt].shape[0], df[mask1 & opt].shape[0], df[mask2 & opt].shape[0]]


def check_va(data, data2):
    
    # IP位址, Service Port, 弱點ID
    total = ((row[0], row[3], row[4]) for row in data)
    total2 = ((row[0], row[3], row[4]) for row in data2)

    set1, set2 = set(total), set(total2)

    not_repair_va = set1 & set2

    new_va = set2 - not_repair_va

    repaired_va = set1 - not_repair_va

    elem = ('IP位址', 'Service Port', '弱點ID')
    new_va.add(elem)
    repaired_va.add(elem)
    
    return list(new_va), list(not_repair_va), list(repaired_va)


def va_to_row(va, data):

    res = []
    tmp_va = copy.deepcopy(va)
    
    for row in data:
        for cod in tmp_va:
            if row[0] == cod[0] and row[3] == cod[1] and row[4] == cod[2]:
                res.append(row)
                tmp_va.remove(cod)
                break
                
    return res


def doc_risk_cnt_compare(doc, data, data2, ips):

    # calculate needed data
    new_va, not_repair_va, repaired_va = check_va(data, data2)

    new_va_page = va_to_row(new_va, data2)
    not_repair_va_page = va_to_row(not_repair_va, data)
    repaired_va_page = va_to_row(repaired_va, data)
    
    excel_risk_cnt = cal_risk_cnt(data, ips)
    excel2_risk_cnt = cal_risk_cnt(not_repair_va_page, ips)
    repaired_risk_cnt = cal_risk_cnt(repaired_va_page, ips)
    
    row_name = ('高風險', '中風險', '低風險')
    col_name = ('', '初掃弱點數', '複掃弱點數', '修補弱點數', '弱點修補率(%)')

    # 替換標定資料
    for table in doc.tables:
        if "riskCntCompare" in table.cell(0, 0).paragraphs[0].text:
            Parag.delete(table.cell(0,0).paragraphs[0])
            table = table.cell(0, 0).add_table(4, 5)
            table.style = 'Table Grid'

            for col in range(5):
                table.cell(0, col).text = col_name[col]
                Table.set_cell_color(table.cell(0, col), GRAY)
                for row in range(1, 4):
                    text = ""
                    if col == 0:
                        text = row_name[row-1]
                        Table.set_cell_color(table.cell(row, 0), GRAY)
                    elif col == 1:
                        text = str(excel_risk_cnt[row-1])
                    elif col == 2:
                        text = str(excel2_risk_cnt[row-1])
                    elif col == 3:
                        text = str(repaired_risk_cnt[row-1])
                    elif col == 4:
                        text = str(100 * repaired_risk_cnt[row-1] // excel_risk_cnt[row-1]) + "%"

                    table.cell(row, col).paragraphs[0].text = text

            Table.set_content_font(table, size=Pt(12))
            

        # row = parag.text
        # if "weaknessCount" in row or "highRiskWeaknessCount" in row or "middleRiskWeaknessCount" in row or "lowRiskWeaknessCount" in row:
        #     row = row.replace("weaknessCount", str(excel2_risk_cnt[0]+excel2_risk_cnt[1]+excel2_risk_cnt[2]), 1)
        #     row = row.replace("highRiskWeaknessCount", str(excel2_risk_cnt[0]), 1)
        #     row = row.replace("middleRiskWeaknessCount", str(excel2_risk_cnt[1]), 1)
        #     row = row.replace("lowRiskWeaknessCount", str(excel2_risk_cnt[2]), 1)
        #     parag.text = row


def compare_out_excel(data, data2, save):

    new_va, not_repair_va, repaired_va = check_va(data, data2)

    new_va_page = va_to_row(new_va, data2)
    not_repair_va_page = va_to_row(not_repair_va, data2)
    repaired_va_page = va_to_row(repaired_va, data)
    
    new_wb = Workbook()

    current_sheet = new_wb.create_sheet('All-初掃', index=0)
    for row in data:
        current_sheet.append(row)

    current_sheet = new_wb.create_sheet('All-複掃', index=1)
    for row in data2:
        current_sheet.append(row)

    current_sheet = new_wb.create_sheet('新掃到弱點', index=2)
    for row in new_va_page:
        current_sheet.append(row)

    current_sheet = new_wb.create_sheet('已修復弱點', index=3)
    for row in repaired_va_page:
        current_sheet.append(row)

    current_sheet = new_wb.create_sheet('未修復弱點', index=4)
    for row in not_repair_va_page:
        current_sheet.append(row)

    new_wb.save(save)


def ExcelToWord_compare(word, excel, excel2, sheet, sheet2, save):

    wb = load_workbook(excel)
    wb2 = load_workbook(excel2)
    data, data2 = list(wb[sheet].values), list(wb2[sheet2].values)
    
    doc = Document(word)
    
    doc_risk_cnt_compare(doc, data, data2, [])

    doc.save(save)


def ExcelToExcel_compare(excel, excel2, sheet, sheet2, save):

    wb = load_workbook(excel)
    wb2 = load_workbook(excel2)
    data, data2 = list(wb[sheet].values), list(wb2[sheet2].values)

    compare_out_excel(data, data2, save)


if __name__ == '__main__':
    
    word = '/Volumes/DATA/GoodleDrive/Internship/公司資料/模板/proto_compare.docx'
    excel = '/Volumes/DATA/GoodleDrive/Internship/公司資料/初掃 xlsx/1216 全聯初掃報告.xlsx'
    excel2 = '/Volumes/DATA/GoodleDrive/Internship/公司資料/複掃 xlsx/0207全聯複掃報告.xlsx'
    sheet = 'All'
    sheet2 = 'All'
    save = '/Users/yucc/Desktop/' + 'demo.docx'
    
    ExcelToWord_compare(word, excel, excel2, sheet, sheet2, save)
